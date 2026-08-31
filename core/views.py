from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.core.cache import cache

from .models import (
    Publication, DeleteRequest, ActivityLog,
    PublicationTypeDict, CitationDatabaseDict, PublicationScopeDict,
    AuthorStatusDict, ReportingPeriodDict, ResultDict,
)
from .serializers import (
    PublicationListSerializer, PublicationDetailSerializer,
    PublicationCreateSerializer, PublicationUpdateSerializer,
    PublicationModerateSerializer,
    DeleteRequestSerializer, DeleteRequestCreateSerializer,
    ActivityLogSerializer
)
from .permissions import (
    CanCreatePublication, CanUpdateOwnPublication, CanDeletePublication,
    CanManageDeleteRequests, CanViewLogs, is_admin, is_methodist, is_nio_staff
)
from users.models import Department


PUBLICATION_STATUS_CHOICES = dict(Publication.STATUS_CHOICES)
MODERATION_STATUS_CHOICES = dict(Publication.MODERATION_STATUS)


@api_view(['GET'])
def reference_data(request):
    """Справочные данные для форм и фильтров."""
    def _data(model):
        return list(model.objects.values('code', 'label'))

    departments = [
        {'code': d.code, 'label': d.full_name}
        for d in Department.objects.all().order_by('code')
    ]

    return Response({
        'departments': departments,
        'publication_types': _data(PublicationTypeDict),
        'citation_databases': _data(CitationDatabaseDict),
        'publication_scopes': _data(PublicationScopeDict),
        'author_statuses': _data(AuthorStatusDict),
        'reporting_periods': _data(ReportingPeriodDict),
        'results': _data(ResultDict),
        'months': [{'code': i, 'label': label} for i, label in Publication.MONTH_CHOICES],
    })


class PublicationViewSet(viewsets.ModelViewSet):
    queryset = Publication.objects.select_related(
        'department', 'publication_type', 'citation_db', 'publication_scope',
        'author_status', 'reporting_period', 'result', 'publisher',
        'owner', 'moderated_by'
    )
    permission_classes = [IsAuthenticated]
    search_fields = ['title', 'author', 'head', 'executors', 'event_name']
    ordering_fields = ['created_at', 'year', 'title']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return PublicationListSerializer
        if self.action == 'retrieve':
            return PublicationDetailSerializer
        if self.action == 'create':
            return PublicationCreateSerializer
        if self.action in ['update', 'partial_update']:
            return PublicationUpdateSerializer
        return PublicationListSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'statistics', 'export']:
            return [AllowAny()]
        if self.action == 'create':
            return [CanCreatePublication()]
        if self.action in ['update', 'partial_update']:
            return [CanUpdateOwnPublication()]
        if self.action == 'destroy':
            return [CanDeletePublication()]
        return [IsAuthenticated()]

    def get_queryset(self):
        # Действия модерации и архивирования работают по полной выборке,
        # права проверяются отдельно внутри действий.
        if self.action in ('moderate', 'toggle_archive', 'restore', 'hard_delete'):
            return super().get_queryset()

        user = self.request.user
        q = super().get_queryset()

        include_archived = self.request.query_params.get('include_archived', 'false').lower() == 'true'

        if not user.is_authenticated:
            q = q.filter(moderation_status='approved', status='active', is_archived=False)
        elif is_admin(user) or is_nio_staff(user):
            if not include_archived:
                q = q.filter(status='active')
        else:
            if not include_archived:
                q = q.filter(
                    Q(moderation_status='approved') | Q(owner=user),
                    status='active',
                    is_archived=False
                )
            else:
                q = q.filter(
                    Q(moderation_status='approved') | Q(owner=user),
                    status='active'
                )

        moderation_status = self.request.query_params.get('moderation_status')
        if moderation_status:
            q = q.filter(moderation_status=moderation_status)

        citation_db = self.request.query_params.get('citation_database')
        if citation_db:
            q = q.filter(citation_db__code=citation_db)

        publication_type = self.request.query_params.get('publication_type')
        if publication_type:
            q = q.filter(publication_type__code=publication_type)

        reporting_period = self.request.query_params.get('reporting_period')
        if reporting_period:
            q = q.filter(reporting_period__code=reporting_period)

        publication_scope = self.request.query_params.get('publication_scope')
        if publication_scope:
            q = q.filter(publication_scope__code=publication_scope)

        result = self.request.query_params.get('result')
        if result:
            q = q.filter(result__code=result)

        department = self.request.query_params.get('department')
        if department:
            q = q.filter(department__code=department)

        year = self.request.query_params.get('year')
        if year:
            try:
                q = q.filter(year=int(year))
            except (ValueError, TypeError):
                pass

        year_from = self.request.query_params.get('year_from')
        if year_from:
            try:
                q = q.filter(year__gte=int(year_from))
            except (ValueError, TypeError):
                pass

        year_to = self.request.query_params.get('year_to')
        if year_to:
            try:
                q = q.filter(year__lte=int(year_to))
            except (ValueError, TypeError):
                pass

        sort_by = self.request.query_params.get('sort_by')
        if sort_by:
            if sort_by == 'year_desc':
                q = q.order_by('-year')
            elif sort_by == 'year_asc':
                q = q.order_by('year')
            elif sort_by == 'title':
                q = q.order_by('title')
            elif sort_by == 'created_desc':
                q = q.order_by('-created_at')
            elif sort_by == 'created_asc':
                q = q.order_by('created_at')

        return q

    def _client_ip(self):
        ip = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if ip:
            return ip.split(',')[0].strip()
        return self.request.META.get('REMOTE_ADDR')

    def perform_create(self, serializer):
        publication = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action='create',
            publication=publication,
            details={'title': publication.title},
            ip_address=self._client_ip(),
        )

    def perform_update(self, serializer):
        publication = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action='update',
            publication=publication,
            details={
                'title': publication.title,
                'changed_fields': list(serializer.validated_data.keys())
            },
            ip_address=self._client_ip(),
        )

    def perform_destroy(self, instance):
        instance.status = 'archived'
        instance.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action='soft_delete',
            publication=instance,
            details={'title': instance.title},
            ip_address=self._client_ip(),
        )

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def moderate(self, request, pk=None):
        publication = self.get_object()
        if not is_admin(request.user) and not is_methodist(request.user) and not is_nio_staff(request.user):
            return Response(
                {'error': 'У вас нет прав для модерации'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = PublicationModerateSerializer(data=request.data)
        if serializer.is_valid():
            old_status = publication.moderation_status
            publication.moderation_status = serializer.validated_data['status']
            publication.moderated_by = request.user
            publication.moderated_at = timezone.now()
            publication.moderation_comment = serializer.validated_data.get('comment', '')
            publication.save()

            action = 'moderation_approved' if publication.moderation_status == 'approved' else 'moderation_rejected'
            ActivityLog.objects.create(
                user=request.user,
                action=action,
                publication=publication,
                details={
                    'old_status': old_status,
                    'new_status': publication.moderation_status,
                    'comment': publication.moderation_comment
                },
                ip_address=self._client_ip(),
            )

            self._send_moderation_notification(publication)

            return Response(PublicationDetailSerializer(publication).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def toggle_archive(self, request, pk=None):
        publication = self.get_object()
        if not is_admin(request.user) and not is_methodist(request.user) and not is_nio_staff(request.user):
            return Response(
                {'error': 'У вас нет прав для архивирования'},
                status=status.HTTP_403_FORBIDDEN
            )

        publication.is_archived = not publication.is_archived
        publication.save()

        return Response({
            'is_archived': publication.is_archived,
            'message': 'Публикация архивирована' if publication.is_archived else 'Публикация восстановлена из архива'
        })

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def archive_old_publications(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Только администратор может архивировать старые публикации'},
                status=status.HTTP_403_FORBIDDEN
            )

        year_limit = 2016
        updated_count = Publication.objects.filter(year__lte=year_limit, is_archived=False).update(is_archived=True)

        return Response({
            'message': f'Архивировано публикаций до {year_limit}: {updated_count}',
            'updated_count': updated_count
        })

    def _send_moderation_notification(self, publication):
        try:
            from django.core.mail import send_mail
            from django.conf import settings

            if not publication.owner or not publication.owner.email:
                return

            if publication.moderation_status == 'approved':
                subject = 'Ваша публикация одобрена'
                message = f'Публикация "{publication.title}" была одобрена.'
            elif publication.moderation_status == 'rejected':
                subject = 'Ваша публикация отклонена'
                message = f'Публикация "{publication.title}" была отклонена.'
                if publication.moderation_comment:
                    message += f'\nКомментарий: {publication.moderation_comment}'
            else:
                return

            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else None,
                recipient_list=[publication.owner.email],
                fail_silently=True,
            )
        except Exception:
            pass

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def my_publications(self, request):
        user = request.user
        queryset = Publication.objects.filter(owner=user, status='active').select_related(
            'department', 'publication_type', 'citation_db', 'result', 'moderated_by', 'owner'
        )

        include_archived = request.query_params.get('include_archived', 'false').lower() == 'true'
        if not include_archived:
            queryset = queryset.filter(is_archived=False)

        queryset = queryset.order_by('-created_at')
        page = self.paginate_queryset(queryset)

        if page is not None:
            serializer = PublicationListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = PublicationListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def deleted(self, request):
        user = request.user
        if not (is_admin(user) or is_nio_staff(user)):
            return Response(
                {'error': 'У вас нет прав для просмотра удалённых записей'},
                status=status.HTTP_403_FORBIDDEN
            )
        queryset = super().get_queryset().filter(status='archived').order_by('-created_at')
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = PublicationListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = PublicationListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def restore(self, request, pk=None):
        publication = self.get_object()
        if not (is_admin(request.user) or is_nio_staff(request.user)):
            return Response(
                {'error': 'У вас нет прав для восстановления записей'},
                status=status.HTTP_403_FORBIDDEN
            )
        if publication.status != 'archived':
            return Response(
                {'error': 'Запись не находится в архиве'},
                status=status.HTTP_400_BAD_REQUEST
            )
        publication.status = 'active'
        publication.is_archived = False
        publication.save()
        ActivityLog.objects.create(
            user=request.user,
            action='restore',
            publication=publication,
            details={'title': publication.title},
            ip_address=self._client_ip(),
        )
        return Response(PublicationDetailSerializer(publication).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def hard_delete(self, request, pk=None):
        publication = self.get_object()
        if not is_admin(request.user):
            return Response(
                {'error': 'Только администратор может безвозвратно удалять записи'},
                status=status.HTTP_403_FORBIDDEN
            )
        if publication.status != 'archived':
            return Response(
                {'error': 'Запись не находится в архиве. Сначала переместите её в архив.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        title = publication.title
        ActivityLog.objects.create(
            user=request.user,
            action='hard_delete',
            publication=None,
            details={'title': title},
            ip_address=self._client_ip(),
        )
        publication.delete()
        return Response(
            {'message': 'Запись безвозвратно удалена'},
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def statistics(self, request):
        cache_key = 'publication_stats_' + '_'.join(
            f'{k}={v}' for k, v in sorted(request.query_params.items())
        )
        cached_data = cache.get(cache_key)

        if cached_data is not None:
            return Response(cached_data)

        queryset = self.get_queryset()

        total = queryset.count()

        department_map = {d.code: d.full_name for d in Department.objects.all()}
        publication_type_map = {d.code: d.label for d in PublicationTypeDict.objects.all()}
        citation_map = {d.code: d.label for d in CitationDatabaseDict.objects.all()}
        scope_map = {d.code: d.label for d in PublicationScopeDict.objects.all()}
        author_status_map = {d.code: d.label for d in AuthorStatusDict.objects.all()}
        period_map = {d.code: d.label for d in ReportingPeriodDict.objects.all()}
        result_map = {d.code: d.label for d in ResultDict.objects.all()}

        by_department = [
            {
                'department': row['department__code'],
                'department_display': department_map.get(row['department__code'], ''),
                'count': row['count'],
            }
            for row in queryset.values('department__code').annotate(count=Count('id')).order_by('-count')
        ]

        by_year = list(queryset.values('year').annotate(count=Count('id')).order_by('-year'))

        by_moderation_status = [
            {
                'moderation_status': row['moderation_status'],
                'moderation_status_display': MODERATION_STATUS_CHOICES.get(row['moderation_status'], row['moderation_status']),
                'count': row['count'],
            }
            for row in queryset.values('moderation_status').annotate(count=Count('id'))
        ]

        by_status = [
            {
                'status': row['status'],
                'status_display': PUBLICATION_STATUS_CHOICES.get(row['status'], row['status']),
                'count': row['count'],
            }
            for row in queryset.values('status').annotate(count=Count('id'))
        ]

        by_publication_type = [
            {
                'publication_type': row['publication_type__code'],
                'publication_type_display': publication_type_map.get(row['publication_type__code'], ''),
                'count': row['count'],
            }
            for row in queryset.values('publication_type__code').annotate(count=Count('id')).order_by('-count')
        ]

        by_citation_database = [
            {
                'citation_db': row['citation_db__code'],
                'citation_db_display': citation_map.get(row['citation_db__code'], ''),
                'count': row['count'],
            }
            for row in queryset.values('citation_db__code').annotate(count=Count('id')).order_by('-count')
        ]

        by_publication_scope = [
            {
                'publication_scope': row['publication_scope__code'],
                'publication_scope_display': scope_map.get(row['publication_scope__code'], ''),
                'count': row['count'],
            }
            for row in queryset.values('publication_scope__code').annotate(count=Count('id')).order_by('-count')
        ]

        by_author_status = [
            {
                'author_status': row['author_status__code'],
                'author_status_display': author_status_map.get(row['author_status__code'], ''),
                'count': row['count'],
            }
            for row in queryset.values('author_status__code').annotate(count=Count('id')).order_by('-count')
        ]

        by_result = [
            {
                'result': row['result__code'],
                'result_display': result_map.get(row['result__code'], ''),
                'count': row['count'],
            }
            for row in queryset.values('result__code').annotate(count=Count('id')).order_by('-count')
        ]

        by_reporting_period = [
            {
                'reporting_period': row['reporting_period__code'],
                'reporting_period_display': period_map.get(row['reporting_period__code'], ''),
                'count': row['count'],
                'total_printed_sheets': row['total_printed_sheets'] or 0,
            }
            for row in queryset.values('reporting_period__code').annotate(
                count=Count('id'),
                total_printed_sheets=Sum('printed_sheets')
            ).order_by('-count')
        ]

        key_metrics = {
            'scopus_wos_count': queryset.filter(
                Q(citation_db__code='WOS') | Q(citation_db__code='SCOPUS')
            ).count(),
            'vak_count': queryset.filter(citation_db__code='VAK').count(),
            'rinc_count': queryset.filter(citation_db__code='RINC').count(),
            'student_count': queryset.filter(
                Q(students_count__gt=0) | Q(author__icontains='студент')
            ).count(),
            'total_printed_sheets': queryset.aggregate(
                total=Sum('printed_sheets')
            )['total'] or 0,
        }

        response_data = {
            'total': total,
            'by_department': by_department,
            'by_year': by_year,
            'by_moderation_status': by_moderation_status,
            'by_status': by_status,
            'by_publication_type': by_publication_type,
            'by_citation_database': by_citation_database,
            'by_publication_scope': by_publication_scope,
            'by_author_status': by_author_status,
            'by_result': by_result,
            'by_reporting_period': by_reporting_period,
            'key_metrics': key_metrics,
        }

        cache.set(cache_key, response_data, 3600)

        return Response(response_data)

    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def export(self, request):
        format_type = request.query_params.get('export_format', 'csv')
        queryset = self.get_queryset()

        if format_type == 'csv':
            return self._export_csv(queryset)
        elif format_type == 'xlsx':
            return self._export_xlsx(queryset)
        elif format_type == 'xml':
            return self._export_xml(queryset)
        return Response(
            {'error': 'Неподдерживаемый формат'},
            status=status.HTTP_400_BAD_REQUEST
        )

    EXPORT_FIELDS = [
        ('id', 'ID'),
        ('title', 'Название публикации / мероприятия'),
        ('author', 'Автор(ы)'),
        ('year', 'Год'),
        ('head', 'Руководитель'),
        ('executors', 'Исполнители'),
        ('students_names', 'ФИО студентов'),
        ('students_count', 'Количество студентов'),
        ('department_full', 'Кафедра'),
        ('publication_type_label', 'Тип публикации'),
        ('publication_scope_label', 'Вид публикации'),
        ('author_status_label', 'Статус автора'),
        ('result_label', 'Результат'),
        ('citation_db_label', 'База цитирования'),
        ('publisher_name', 'Издательство'),
        ('location', 'Место проведения'),
        ('event_name', 'Название мероприятия'),
        ('event_date', 'Дата проведения'),
        ('pages_count', 'Количество страниц'),
        ('printed_sheets', 'Печатные листы'),
        ('circulation', 'Тираж'),
        ('volume', 'Объём'),
        ('funding_source', 'Источник финансирования'),
        ('doi', 'DOI'),
        ('edn_code', 'EDN'),
        ('elibrary_id', 'ID eLibrary'),
        ('reporting_period_label', 'Отчётный период'),
        ('reporting_year', 'Отчётный год'),
        ('entry_month', 'Месяц внесения'),
        ('keywords', 'Ключевые слова'),
        ('note', 'Примечание'),
        ('status_display', 'Статус'),
        ('moderation_status', 'Статус модерации'),
        ('moderation_comment', 'Комментарий модерации'),
        ('owner_username', 'Владелец'),
        ('is_archived', 'В архиве'),
        ('created_at', 'Дата создания'),
        ('updated_at', 'Дата обновления'),
    ]

    def _export_row_data(self, p):
        row = {}
        for key, _header in self.EXPORT_FIELDS:
            if key == 'department_full':
                row[key] = p.department.full_name if p.department else ''
            elif key == 'publication_type_label':
                row[key] = p.publication_type.label if p.publication_type else ''
            elif key == 'publication_scope_label':
                row[key] = p.publication_scope.label if p.publication_scope else ''
            elif key == 'author_status_label':
                row[key] = p.author_status.label if p.author_status else ''
            elif key == 'result_label':
                row[key] = p.result.label if p.result else ''
            elif key == 'citation_db_label':
                row[key] = p.citation_db.label if p.citation_db else ''
            elif key == 'publisher_name':
                row[key] = p.publisher.name if p.publisher else ''
            elif key == 'reporting_period_label':
                row[key] = p.reporting_period.label if p.reporting_period else ''
            elif key == 'status_display':
                row[key] = p.get_status_display()
            elif key == 'owner_username':
                row[key] = p.owner.username if p.owner else ''
            elif key == 'is_archived':
                row[key] = 'Да' if p.is_archived else 'Нет'
            elif key == 'entry_month':
                row[key] = p.get_entry_month_display() if hasattr(p, 'get_entry_month_display') else p.entry_month
            elif key in ('created_at', 'updated_at'):
                row[key] = p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else ''
            elif key == 'event_date':
                row[key] = p.event_date.strftime('%Y-%m-%d') if p.event_date else ''
            else:
                value = getattr(p, key)
                row[key] = value if value is not None else ''
        return row

    def _export_rows(self, queryset):
        from django.db.models import F
        qs = queryset.select_related(
            'department', 'publication_type', 'publication_scope',
            'author_status', 'result', 'citation_db', 'reporting_period',
            'publisher', 'owner'
        )
        headers = [h for _k, h in self.EXPORT_FIELDS]
        data_rows = [self._export_row_data(p) for p in qs]
        return headers, data_rows

    def _export_csv(self, queryset):
        import csv
        from django.http import HttpResponse
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)
        headers, data_rows = self._export_rows(queryset)
        writer.writerow(headers)
        for d in data_rows:
            writer.writerow([d[k] for k, _h in self.EXPORT_FIELDS])

        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="publications.csv"'
        return response

    def _export_xlsx(self, queryset):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'openpyxl не установлен'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        from django.http import HttpResponse
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Публикации'

        headers, data_rows = self._export_rows(queryset)

        header_fill = PatternFill(start_color='1A6F3F', end_color='1A6F3F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        for ridx, d in enumerate(data_rows, 2):
            for cidx, (key, _h) in enumerate(self.EXPORT_FIELDS, 1):
                cell = ws.cell(row=ridx, column=cidx, value=d[key])
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        col_widths = {
            'title': 60, 'author': 40, 'executors': 40, 'students_names': 40,
            'note': 40, 'keywords': 30, 'moderation_comment': 30, 'event_name': 40,
            'location': 30, 'funding_source': 30, 'volume': 20,
        }
        for cidx, (key, _h) in enumerate(self.EXPORT_FIELDS, 1):
            width = col_widths.get(key, 22)
            ws.column_dimensions[get_column_letter(cidx)].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="publications.xlsx"'
        return response

    def _export_xml(self, queryset):
        from django.http import HttpResponse
        import xml.etree.ElementTree as ET

        headers, data_rows = self._export_rows(queryset)
        root = ET.Element('publications')
        for d in data_rows:
            pub = ET.SubElement(root, 'publication')
            for key, _h in self.EXPORT_FIELDS:
                ET.SubElement(pub, key).text = str(d[key])

        xml_str = ET.tostring(root, encoding='unicode')
        response = HttpResponse(xml_str, content_type='application/xml')
        response['Content-Disposition'] = 'attachment; filename="publications.xml"'
        return response


class DeleteRequestViewSet(viewsets.ModelViewSet):
    queryset = DeleteRequest.objects.select_related('publication', 'requester', 'reviewed_by')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return DeleteRequestCreateSerializer
        return DeleteRequestSerializer

    def get_permissions(self):
        if self.action in ['update', 'partial_update', 'destroy']:
            return [CanManageDeleteRequests()]
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        if is_admin(user):
            return self.queryset.all()
        return self.queryset.filter(requester=user)

    def perform_update(self, serializer):
        instance = serializer.save()
        instance.reviewed_by = self.request.user
        instance.reviewed_at = timezone.now()
        instance.save()

        action = 'delete_request_approved' if instance.status == 'approved' else 'delete_request_rejected'
        ActivityLog.objects.create(
            user=self.request.user,
            action=action,
            publication=instance.publication,
            details={'request_id': instance.id}
        )


class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ActivityLog.objects.select_related('user', 'publication')
    permission_classes = [CanViewLogs]
    serializer_class = ActivityLogSerializer
    filter_fields = ['action', 'user']
    ordering_fields = ['timestamp']
    ordering = ['-timestamp']

    def get_queryset(self):
        user = self.request.user
        if is_admin(user):
            return self.queryset.all()
        return ActivityLog.objects.none()